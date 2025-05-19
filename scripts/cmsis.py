#!/usr/bin/env python3
import openpyxl
import requests
import tarfile
import json
from hashlib import sha256
from dataclasses import dataclass, asdict, is_dataclass
from pprint import pprint
from collections import defaultdict
from pathlib import Path
from packaging.version import Version
import typing as t
from jinja2 import Environment, FileSystemLoader
import re

CORES = {'cortex-m0', 'cortex-m0plus', 'cortex-m33', 'cortex-m3', 'cortex-m4', 'cortex-m7', 'cortex-m55'}

@dataclass
class SysOpt:
    name: str
    type: str
    description: str
    define: str

@dataclass
class PkgInfo:
    name: str
    version: str
    tarball_url: str
    tarball_sha256: str
    root_dir: str
    mcu_types: list
    system_options: t.List[SysOpt]
    mcu_specs: list

    def __init__(self, name, version=None):
        self.name = name
        self.version = version
        self.tarball_url = None
        self.tarball_sha256 = None
        self.root_dir = None
        self.mcu_types = None
        self.system_options = None
        self.mcu_specs = []
        self.out_path = None
        self.patch_dir = None

    def set_out_path(self, out_path):
        self.out_path = out_path
        self.patch_dir = out_path + f'/packagefiles/st-cmsis-{self.name}'

    def set_mcu_types(self, tarball_content, tb):
        self.mcu_types = []
        root = tarball_content[0]

        if self.name == 'wl3':
            main_include = None
        else:
            main_include = next(filter(lambda x: x==root+'/Include/stm32'+self.name+'xx.h', tarball_content), None)

        if main_include is None:
            # some quirks
            if self.name == 'wb0':
                self.mcu_types.extend(['stm32wb0_', 'stm32wb05', 'stm32wb06', 'stm32wb07', 'stm32wb09'])
            if self.name == 'wl3':
                self.mcu_types.append('stm32wl3__')
        else:
            inc = tb.extractfile(main_include)
            prev_line = None
            for l in inc:
                if l.strip().startswith(b'#include'):
                    h = l.strip()
                    d = prev_line.strip().replace(b'#if defined', b'').replace(b'#elif defined', b'').replace(b'#ifdef', b'')
                    d = d.replace(b'(', b'').replace(b')', b'').strip()
                    if d.startswith(b'STM32'):
                        self.mcu_types.append(d.decode('ascii').replace('x', '_').lower())
                prev_line = l

        if len(self.mcu_types) == 0:
            raise RuntimeError('Cannot detect types for ' + self.name)

    def create_patch_dir(self):
        Path(self.patch_dir).mkdir(parents=True, exist_ok=True)

    def create_wrap(self, temp):
        name = f'/st-cmsis-{self.name}.wrap'
        with open(self.out_path + name, 'w') as f:
            f.write(temp.render(asdict(self)))

    def create_build(self, temp):
        with open(self.patch_dir + '/meson.build', 'w') as f:
            f.write(temp.render(asdict(self)))

    def create_options(self, temp):
        with open(self.patch_dir + '/meson.options', 'w') as f:
            f.write(temp.render(asdict(self)))

    def create_ld(self, temp):
        ld_dir = Path(self.patch_dir) / 'ld'
        ld_dir.mkdir(parents=True, exist_ok=True)
        for spec in self.mcu_specs:
            mcu, flash_kb, ram_kb, ccmram_kb = spec[0], spec[2], spec[3], spec[4]
            with open(ld_dir / f'{mcu}.ld', 'w') as f:
                f.write(temp.render({'mcu': mcu, 'flash_kb': flash_kb, 'ram_kb': ram_kb, 'ccmram_kb': ccmram_kb}))

PACKAGES = [
    PkgInfo('c0', '1.3.0'),
    PkgInfo('f0', '2.3.7'),
    PkgInfo('f1', '4.3.5'),
    PkgInfo('f2', '2.2.6'),
    PkgInfo('f3', '2.3.8'),
    PkgInfo('f4', '2.6.10'),
    PkgInfo('f7', '1.2.9'),
    PkgInfo('g0', '1.4.4'),
    PkgInfo('g4', '1.2.5'),
    PkgInfo('h5', '1.4.0'),
    PkgInfo('h7', '1.10.6'),
    PkgInfo('h7rs', '1.2.0'),
    PkgInfo('l0', '1.9.4'),
    PkgInfo('l1', '2.3.4'),
    PkgInfo('l4', '1.7.4'),
    PkgInfo('l5', '1.0.6'),
    PkgInfo('n6', '1.1.0'),
    PkgInfo('u0', '1.2.0'),
    PkgInfo('u3', '1.1.0'),
    PkgInfo('u5', '1.4.1'),
    PkgInfo('wb', '1.12.2'),
    PkgInfo('wb0', '1.2.0'),
    PkgInfo('wba', '1.6.0'),
    PkgInfo('wl', '1.2.0'),
    PkgInfo('wl3', '1.1.0'),
]

HSE_OPT = SysOpt('hse_value', 'integer', 'Value of the External oscillator in Hz', 'HSE_VALUE')
LSE_OPT = SysOpt('lse_value', 'integer', 'Value of the External oscillator in Hz', 'LSE_VALUE')
HSI_OPT = SysOpt('hsi_value', 'integer', 'Value of the Internal oscillator in Hz', 'HSI_VALUE')
MSI_OPT = SysOpt('msi_value', 'integer', 'Value of the Internal oscillator in Hz', 'MSI_VALUE')
LSI_OPT = SysOpt('lsi_value', 'integer', 'Value of the Internal oscillator in Hz', 'LSI_VALUE')
HSI48_OPT = SysOpt('hsi48_value', 'integer', 'Value of the HSI48 oscillator in Hz', 'HSI48_VALUE')
VTAB_SRAM_OPT = SysOpt('vect_tab_sram', 'boolean', 'Relocate vector Table in Internal SRAM', 'VECT_TAB_SRAM')
USE_VTOR_OPT = SysOpt('user_vect_tab_address', 'boolean', 'Relocate the vector table anywhere in Flash or SRAM', 'USER_VECT_TAB_ADDRESS')
VTAB_OFFSET_OPT = SysOpt('vect_tab_offset', 'string', 'Hex with 0x prefix. This value must be a multiple of 0x200.', 'VECT_TAB_OFFSET')
DATA_EXT_SRAM_OPT = SysOpt('data_in_ext_sram', 'boolean', 'Use external SRAM', 'DATA_IN_ExtSRAM')
DATA_EXT_SDRAM_OPT = SysOpt('data_in_ext_sdram', 'boolean', 'Use external SDRAM', 'DATA_IN_ExtSDRAM')
SYSTEM_OPTIONS = {
    'c0': [HSE_OPT, HSI_OPT, LSI_OPT, LSE_OPT, HSI48_OPT, VTAB_SRAM_OPT],
    'f0': [HSE_OPT, HSI_OPT, HSI48_OPT],
    'f1': [HSE_OPT, HSI_OPT, DATA_EXT_SRAM_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'f2': [HSE_OPT, HSI_OPT, DATA_EXT_SRAM_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'f3': [HSE_OPT, HSI_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'f4': [HSE_OPT, HSI_OPT, DATA_EXT_SRAM_OPT, DATA_EXT_SDRAM_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'f7': [HSE_OPT, HSI_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'g0': [HSE_OPT, HSI_OPT, LSI_OPT, LSE_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'g4': [HSE_OPT, HSI_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'l0': [HSE_OPT, HSI_OPT, MSI_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'l1': [HSE_OPT, HSI_OPT, DATA_EXT_SRAM_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT, VTAB_OFFSET_OPT],
    'l4': [HSE_OPT, HSI_OPT, MSI_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
    'u0': [HSE_OPT, HSI_OPT, LSE_OPT, MSI_OPT, LSI_OPT, VTAB_SRAM_OPT],
    'wb': [HSE_OPT, HSI_OPT, LSE_OPT, MSI_OPT, LSI_OPT, USE_VTOR_OPT, VTAB_SRAM_OPT],
}

@dataclass
class McuInfo:
    name: str
    cores: str
    st_package: str
    flash_kb: int
    ram_kb: int


def get_github_tags(family):
    url = f"https://api.github.com/repos/STMicroelectronics/cmsis-device-{family}/tags"
    resp = requests.get(url).json()
    assert 'message' not in resp, resp
    return resp

def latest_tag(tags):
    latest_tag = tags[0]
    for tag in tags:
        if Version(latest_tag['name']) < Version(tag['name']):
            latest_tag = tag
    return latest_tag

def get_tarball(url):
    return requests.get(url).content



GARBAGE = [
    ('stm32c', ['n', 'ntr']),
    ('stm32f1', ['a', 'b', 'v', 'w', 'btr', 'atr']),
    ('stm32f2', ['v', 'w']),
    ('stm32f4', ['v', 'w']),
    ('stm32f', ['i', 'tt', 'vtr', 'wtr', 'ptr', 'mtr']),
    ('stm32g', ['n', 'ntr']),
    ('stm32l0', ['d', 's', 'dtr', 'str']),
    ('stm32l1', ['d', 'dtr', 'ttr']),
    ('stm32l4', ['s', 'p', 'str', 'mtr', 'ptr', 'pst', 'f']),
    ('stm32l5', ['q', 'p', 'qtr', 'ptr']),
    ('stm32h5', ['q', 'qtr']),
    ('stm32h7', ['a', 'h']),
    ('stm32n6', ['q', 'qtr', 'qs']),
    ('stm32u3', ['q', 'qtr', 'gtr']),
    ('stm32u5', ['q', 'qtr']),
    ('stm32wb1', ['e']),
    ('stm32wb3', ['a', 'atr']),
    ('stm32wl3', ['a', 'atr', 'x']),
    ('stm32wl5', ['str']),
    ('', ['tr']),
]
def remove_garbage(pnum):
    for prfx, sfx in GARBAGE:
        if pnum.startswith(prfx):
            for s in sfx:
                if pnum.endswith(s):
                    pnum = pnum[:-len(s)]
    return pnum

def parse_products_list(path):
    mcus = list()
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    col_names = tuple(x.value for x in [x for x in sheet.iter_rows(10, 10)][0])
    already = dict()
    for i in range(11, sheet.max_row + 1):
        info = McuInfo(None, None, None, None, None)
        values = dict()
        for c, name in enumerate(col_names):
            values[name] = sheet.cell(row=i, column=c+1).value

        pnum = values['Part Number'].lower()
        pnum = remove_garbage(pnum)
        idx = -1
        internal_code = None
        if pnum.startswith('stm32l1') and pnum[-1] in 'ax':
            internal_code = pnum[-1]
            idx = -2
        if pnum.startswith('stm32h7') and pnum[-1] in 'q':
            internal_code = pnum[-1]
            idx = -2
        if not pnum[idx].isdigit():
            raise RuntimeError(pnum, len(pnum))
        pnum = pnum[:-2] # remove package and temperature range
        if len(pnum) != 11:
            # 3 symbols for Product Type
            if pnum.startswith('stm32h7') or pnum.startswith('stm32l1'):
                pnum = pnum[:-1]
            elif pnum.startswith('stm32wba'):
                pass
            else:
                raise RuntimeError('Wrong len', pnum, len(pnum))
        if internal_code:
            pnum += internal_code
        values.pop('Part Number')
        if pnum in already:
            if already[pnum] != values:
                raise RuntimeError('somthing wrong ', pnum, values, already[pnum])
            continue
        else:
            info.name = pnum
            already[pnum] = values

        normalized_cores = values['Core'].lower().replace('arm ', '').replace('+', 'plus').split(', ')
        if info.name.startswith('stm32wb05') and info.name[10] == 'n':
            normalized_cores = ['cortex-m0plus']
        for core in normalized_cores:
            if core not in CORES:
                raise RuntimeError(f'invalid core: {info.name}, {normalized_cores}')
        info.cores = normalized_cores

        if info.name.startswith('stm32h7r') or info.name.startswith('stm32h7s'):
            info.st_package = 'h7rs'
        else:
            best_match = ''
            for pkg in PACKAGES:
                if info.name[5:5+len(pkg.name)] == pkg.name and len(pkg.name) > len(best_match):
                    best_match = pkg.name
            if best_match == '':
                raise RuntimeError('Cannot detect package ', pnum)
            info.st_package = best_match

        if values['Flash Size (kB) (Prog)'] == '-':
            print('flash size is - ',info.name)
            info.flash_kb = 0
        else:
            info.flash_kb = int(values['Flash Size (kB) (Prog)'])
        if values['RAM Size (kB)'] == '-':
            print('ram size is - ', info.name)
            info.ram_kb = 0
        else:
            info.ram_kb = int(values['RAM Size (kB)'])

        mcus.append(info)

    return mcus


def enrich_pkg_info(info, tarball_cache):
    if info.version is None:
        tag = latest_tag(get_github_tags(info))
        info.version = tag['name'][1:]  # remove 'v' symbol

    info.tarball_url = f'https://github.com/STMicroelectronics/cmsis-device-{info.name}/archive/refs/tags/v{info.version}.tar.gz'
    tarball_name = f'{info.name}-{info.version}.tar.gz'
    path = Path(tarball_cache, tarball_name)
    if not path.exists():
        tb_bin = get_tarball(info.tarball_url)
        with open(path, 'wb') as f:
            f.write(tb_bin)
    else:
        with open(path, 'rb') as f:
            tb_bin = f.read()

    h1 = sha256()
    h1.update(tb_bin)
    info.tarball_sha256 = h1.hexdigest()

    tb = tarfile.open(path)
    content = tb.getnames()

    info.root_dir = content[0]
    info.set_mcu_types(content, tb)
    info.system_options = None

MANUAL_TYPES = {
    'stm32f030f4': 'stm32f030_6',
    'stm32f100c4': 'stm32f100_b',
    'stm32f100c6': 'stm32f100_b',
    'stm32f100c8': 'stm32f100_b',
    'stm32f100r4': 'stm32f100_b',
    'stm32f100r6': 'stm32f100_b',
    'stm32f100r8': 'stm32f100_b',
    'stm32f100v8': 'stm32f100_b',
    'stm32f100rc': 'stm32f100_e',
    'stm32f100rd': 'stm32f100_e',
    'stm32f100vc': 'stm32f100_e',
    'stm32f100vd': 'stm32f100_e',
    'stm32f100zc': 'stm32f100_e',
    'stm32f100zd': 'stm32f100_e',
    'stm32f101c4': 'stm32f101_6',
    'stm32f101c8': 'stm32f101_b',
    'stm32f101r4': 'stm32f101_6',
    'stm32f101r8': 'stm32f101_b',
    'stm32f101rc': 'stm32f101_e',
    'stm32f101rd': 'stm32f101_e',
    'stm32f101rf': 'stm32f101_g',
    'stm32f101t4': 'stm32f101_6',
    'stm32f101t8': 'stm32f101_b',
    'stm32f101v8': 'stm32f101_b',
    'stm32f101vc': 'stm32f101_e',
    'stm32f101vd': 'stm32f101_e',
    'stm32f101vf': 'stm32f101_g',
    'stm32f101zc': 'stm32f101_e',
    'stm32f101zd': 'stm32f101_e',
    'stm32f101zf': 'stm32f101_g',
    'stm32f102c4': 'stm32f102_6',
    'stm32f102c8': 'stm32f102_b',
    'stm32f102r4': 'stm32f102_6',
    'stm32f102r8': 'stm32f102_b',
    'stm32f103c4': 'stm32f103_6',
    'stm32f103c8': 'stm32f103_b',
    'stm32f103r4': 'stm32f103_6',
    'stm32f103r8': 'stm32f103_b',
    'stm32f103rc': 'stm32f103_e',
    'stm32f103rd': 'stm32f103_e',
    'stm32f103rf': 'stm32f103_g',
    'stm32f103t4': 'stm32f103_6',
    'stm32f103t8': 'stm32f103_b',
    'stm32f103v8': 'stm32f103_b',
    'stm32f103vc': 'stm32f103_e',
    'stm32f103vd': 'stm32f103_e',
    'stm32f103vf': 'stm32f103_g',
    'stm32f103zc': 'stm32f103_e',
    'stm32f103zd': 'stm32f103_e',
    'stm32f103zf': 'stm32f103_g',
    'stm32f302c6': 'stm32f302_8',
    'stm32f302cb': 'stm32f302_c',
    'stm32f302k6': 'stm32f302_8',
    'stm32f302r6': 'stm32f302_8',
    'stm32f302rb': 'stm32f302_c',
    'stm32f302rd': 'stm32f302_e',
    'stm32f302vb': 'stm32f302_c',
    'stm32f302vd': 'stm32f302_e',
    'stm32f302zd': 'stm32f302_e',
    'stm32f303c6': 'stm32f303_8',
    'stm32f303cb': 'stm32f303_c',
    'stm32f303k6': 'stm32f303_8',
    'stm32f303r6': 'stm32f303_8',
    'stm32f303rb': 'stm32f303_c',
    'stm32f303rd': 'stm32f303_e',
    'stm32f303vb': 'stm32f303_c',
    'stm32f303vd': 'stm32f303_e',
    'stm32f303zd': 'stm32f303_e',
    'stm32f401cb': 'stm32f401_c',
    'stm32f401cd': 'stm32f401_e',
    'stm32f401rb': 'stm32f401_c',
    'stm32f401rd': 'stm32f401_e',
    'stm32f401vb': 'stm32f401_c',
    'stm32f401vd': 'stm32f401_e',
    'stm32f778ai': 'stm32f779__',
    'stm32h745bi': 'stm32h745__',
    'stm32h745ii': 'stm32h745__',
    'stm32h745xi': 'stm32h745__',
    'stm32h745zi': 'stm32h745__',
    'stm32h747ai': 'stm32h747__',
    'stm32h747bi': 'stm32h747__',
    'stm32h747ii': 'stm32h747__',
    'stm32h747xi': 'stm32h747__',
    'stm32h747zi': 'stm32h747__',
    'stm32l151c6a': 'stm32l151_ba',
    'stm32l151c8a': 'stm32l151_ba',
    'stm32l151r6a': 'stm32l151_ba',
    'stm32l151r8a': 'stm32l151_ba',
    'stm32l151v8a': 'stm32l151_ba',
    'stm32l152c6a': 'stm32l152_ba',
    'stm32l152c8a': 'stm32l152_ba',
    'stm32l152r6a': 'stm32l152_ba',
    'stm32l152r8a': 'stm32l152_ba',
    'stm32l152v8a': 'stm32l152_ba',
    'stm32wb05kn': 'stm32wb0_',
    'stm32wb05kz': 'stm32wb0_',
    'stm32wb05tn': 'stm32wb0_',
    'stm32wb05tz': 'stm32wb0_',
    'stm32wb06cc': 'stm32wb0_',
    'stm32wb06kc': 'stm32wb0_',
    'stm32wb07cc': 'stm32wb0_',
    'stm32wb07kc': 'stm32wb0_',
    'stm32wb09ke': 'stm32wb0_',
    'stm32wb09te': 'stm32wb0_',
    'stm32l100c6': 'stm32l100_b',
    'stm32l100r8': 'stm32l100_b',
    'stm32l151c6': 'stm32l151_b',
    'stm32l151c8': 'stm32l151_b',
    'stm32l151r6': 'stm32l151_b',
    'stm32l151r8': 'stm32l151_b',
    'stm32l151v8': 'stm32l151_b',
    'stm32l152c6': 'stm32l152_b',
    'stm32l152c8': 'stm32l152_b',
    'stm32l152r6': 'stm32l152_b',
    'stm32l152r8': 'stm32l152_b',
    'stm32l152v8': 'stm32l152_b',
}

ccmram_re = {
    'f40[57]..': 64,
    'f41[57]..': 64,
    'f42[79]..': 64,
    'f43[79]..': 64,
    'f4[67]9..': 64,
}

def main():
    pkgs = {}
    for pkg in PACKAGES:
        enrich_pkg_info(pkg, 'tmp')
        pkgs[pkg.name] = pkg
    mcus = parse_products_list('ProductsList.xlsx')
    # pprint(mcus)
    cpus = dict()
    for mcu in mcus:
        pkg = pkgs[mcu.st_package]
        scores = []
        if mcu.name in MANUAL_TYPES:
            mcu_type = MANUAL_TYPES[mcu.name]
        else:
            for try_mcu_type in pkg.mcu_types:
                dist = 0
                if len(try_mcu_type) == len(mcu.name):
                    for i in range(5, len(mcu.name)):
                        if try_mcu_type[i] == mcu.name[i]:
                            dist |= 1
                        dist <<= 1
                scores.append((dist, try_mcu_type))
            scores.sort(reverse=True)
            if len(scores)!=1 and scores[0][0] == scores[1][0]:
                # best = []
                # for s in scores:
                #     if s[0] == scores[0][0]:
                #         best.append(s)
                # print(f"    '{mcu.name}': {[x[1] for x in best]},")
                raise RuntimeError(mcu.name, scores)
            mcu_type = scores[0][1]

        ccmram_kb = 0
        for k, v in ccmram_re.items():
            if re.match(k, mcu.name[5:12]):
                ccmram_kb = v
        pkg.mcu_specs.append((mcu.name, mcu_type, mcu.flash_kb, mcu.ram_kb, ccmram_kb))
        if mcu.st_package not in cpus:
            cpus[mcu.st_package]= set()
        cpus[mcu.st_package].add(tuple(mcu.cores))
    pprint(cpus)

    # Generate
    jenv = Environment(loader=FileSystemLoader('templates/'), trim_blocks=True)
    wrap_temp = jenv.get_template('cmsis.wrap')
    meson_options_temp = jenv.get_template('meson.options')
    meson_build_temp = jenv.get_template('meson.build')
    ld_temp = jenv.get_template('default.ld')
    path = 'out/subprojects'

    for pkg in pkgs.values():
        pkg.set_out_path(path)
        pkg.create_patch_dir()
        pkg.create_wrap(wrap_temp)
        pkg.create_build(meson_build_temp)
        pkg.create_options(meson_options_temp)
        pkg.create_ld(ld_temp)

if __name__ == '__main__':
    main()
